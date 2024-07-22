from django.shortcuts import render

from tienda.utils import render_pdf
from .models import Producto
#modifique para el xml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl import Workbook#libreria para xl
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from django.views.generic import TemplateView,View
import requests

from urllib.request import urlopen
import json
"""
class ReporteAutorExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        producto=Producto.objects.all()
        wb=Workbook()
        ws=wb.active
        wb['A1']="Reporte Productos"
        ws.merge_cells('A1:C1')

        ws['A3']='Producto'
        ws['B3']='precio'
        ws['C3']='imagen'

        cont=4

        for product in producto:
            ws.cell(row=cont, column=1).values=product.nombre
            ws.cell(row=cont, column=2).values=product.imagen
            ws.cell(row=cont, column=3).values=product.precio
            cont+=1
        
        nombre_archivo="ReporteProductoExcel.xlsx"
        response=HttpResponse(content_type="application/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response"""

class ReporteAutorExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Productos"
        
        ws['A1'] = "Reporte Productos"
        ws.merge_cells('A1:C1')

        ws['A3'] = 'Producto'
        ws['B3'] = 'Precio'
        ws['C3'] = 'Imagen'

        cont = 4

        for product in producto:
            ws.cell(row=cont, column=1).value = product.nombre
            ws.cell(row=cont, column=2).value = product.precio
            ws.cell(row=cont, column=3).value = product.imagen.url
            cont += 1
        
        nombre_archivo = "ReporteProductoExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



# Create your views here.
def tienda(request):
    produ=Producto.objects.all()
    return render(request,'tienda/tienda.html',{"productos":produ})
#modifique 11 de junio

def api(request):
    url="https://random-data-api.com/api/commerce/random_commerce?size=100"
    response = urlopen(url)
#print(response.read())
    data= json.load(response)
    diccionario={}
    j=0
    for i in data:
        if i["material"] not in diccionario.values():
            diccionario[j]=i["material"]
            j+=1
# print(diccionario)
    return render(request,'tienda/api.html',{"diccionario":diccionario})

def listado(request):
    produ=Producto.objects.all()
    return render(request,'tienda/listado.html',{"productos":produ})

def reporte(request):
    produ=Producto.objects.all()
    return render(request,'tienda/listado.html',{"productos":produ})

class ListaProductoPdf(View):
    def get(self,request,*args,**kwargs):
        productos=Producto.objects.all()
        data={
            'productos':productos
        }
        pdf=render_pdf('tienda/listado.html',data)
        return HttpResponse(pdf,content_type='application/pdf')

def evaluacion(request):


    url = "https://numbers-to-words1.p.rapidapi.com/api/converter/"

    payload = {
	"number": 141.2234,
	"delete_from_sentence": None,
	"currency": "euros",
	"decimal_currency": "millimes",
	"separator": "et",
	"decimal": 3,
	"language": "fr"
    }
    headers = {
	"x-rapidapi-key": "3f8f9c086bmsh2560f35958e2861p187b94jsn45c4186b790d",
	"x-rapidapi-host": "numbers-to-words1.p.rapidapi.com",
	"Content-Type": "application/json"
    }

    response = requests.post(url, json=payload, headers=headers)

# print(response.json())
    data=response.json()
    return render(request,'tienda/listado.html',{"data":data})

